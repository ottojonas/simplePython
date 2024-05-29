# To use this code:

#  1.Replace the placeholder values with your desired URLs, CSS selectors, file paths, and sheet names.
#  2.Set the start_index and skip_value variables if you need to iterate through multiple pages.
#  3.Execute the code to launch the Chrome browser and start scraping the data.
#  4.The scraped data will be stored in the Excel file specified.
#  5.After the scraping is complete, the browser will be closed, and the Excel file will be saved.




import openpyxl
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException

# Launch the Chrome Browser
driver = webdriver.Chrome()

# Start page index and skip value
start_index = 12 # use this if you have a skip value in the url
skip_value = 12 # use this if you have a skip value in the url

# Load existing Excel workbook or create a new one if it doesn't exist
try:
    workbook = openpyxl.load_workbook("E:\\xx.xlsx") # xx = rename xls file
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "xx"  #xx means sheet title

# Find the first empty row in the Excel sheet
empty_row = sheet.max_row + 1

# Write the column headers if the sheet is empty
if empty_row == 2:
    column_names = [
        "xx" #xx means column header name add as many as needed        
    ]
    for column_index, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=column_index).value = column_name

# Scrape data and check xx column
while True:
    # Update the URL with the skip value
    url = f"https://www.thesite.org" #add your site as needed add skip if need may be named differently you should get it

    # Navigate to the website
    driver.get(url)

    # Find all xx.xx elements
    item_elements = driver.find_elements(By.CSS_SELECTOR, "xx.xx")  #xx.xx = css selector

    # If no item elements are found, exit the loop
    if len(item_elements) == 0:
        break

    # Iterate over each item element
    for item_element in item_elements:
        # Find the x[class='x'] element within the current item element
        try:
            first_element = item_element.find_element(By.CSS_SELECTOR, "x[class='x']") #x equals your css and selector
            first_data = first_element.text
        except NoSuchElementException:
            first_data = ""

            # Add the above as needed       

        # Check the first data if it is populuated
        if sheet.max_row > 1:
            sheet.cell(row=empty_row, column=1).value = first_data
            
            empty_row += 1

    # Increment the start index for the next page ** for multiple page scrapes
    start_index += skip_value

# Save the Excel file
workbook.save("E:\\xx.xlsx")


# Close the browser
driver.quit()